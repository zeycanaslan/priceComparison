import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Örnek veri
data = {
    'Item Name': ['Ürün 1', 'Ürün 2', 'Ürün 3'],
    'Item Price': [100, 200, 300],
    'Item Link': ['https://urun1-linki', 'https://urun2-linki', 'https://urun3-linki']
}

# Verileri DataFrame'e dönüştür
df = pd.DataFrame(data)

# Excel dosyası oluşturma
workbook = Workbook()
worksheet = workbook.active

# Başlık satırını ekle
header_font = Font(bold=True)
for col_num, column_name in enumerate(df.columns, start=1):
    cell = worksheet.cell(row=1, column=col_num, value=column_name)
    cell.font = header_font

# Veri satırlarını ekle
for row_num, row_data in enumerate(df.values, start=2):
    item_name = row_data[0]
    item_link = row_data[2]
    
    # Ürün adını hiperlink olarak ekle
    cell = worksheet.cell(row=row_num, column=1, value=item_name)
    cell.hyperlink = item_link
    cell.font = Font(underline='single', color='0563C1')

    # Diğer veri sütunlarını ekle
    for col_num, value in enumerate(row_data[1:], start=2):
        cell = worksheet.cell(row=row_num, column=col_num, value=value)

# Hücre genişliklerini otomatik ayarla
for col_num, column in enumerate(df.columns, start=1):
    column_letter = get_column_letter(col_num)
    max_length = max(df[column].astype(str).map(len).max(), len(column))
    adjusted_width = (max_length + 2) * 1.2
    worksheet.column_dimensions[column_letter].width = adjusted_width

# Excel dosyasını kaydet
workbook.save('tablo.xlsx')
